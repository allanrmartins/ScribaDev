"""Exportação Excel do apontamento de horas (épico #118, #122).

Gera um arquivo NOVO no layout da planilha pessoal de apontamento (colunas A-M,
uma aba por mês) — a planilha original nunca é tocada; se o nome já existe,
sufixa _2, _3... Layout conferido contra a planilha real em 2026-07-13:
cabeçalhos, formatos (A d-mmm, horas h:mm), fórmula D =C-B, marca de hora extra
"hora extra" na coluna H e dropdowns (I: Base/In Loco/Remoto, J: SIM/NÃO).

Generalizações deliberadas sobre o original:
- K (Total dia) sai só na ÚLTIMA linha de cada dia, somando o intervalo exato do
  dia — o original usa janelas fixas de 2 linhas sobrepostas, que só fazem
  sentido em dias com exatamente 2 blocos;
- L/M (Ñ Apontado / Apontados) viram células-resumo do MÊS na linha 2, com
  SUMIF de coluna inteira, em formato [h]:mm (totais de mês passam de 24 h).

Só apontamentos `confirmed` entram; `day_notes` (feriado etc.) viram linha com
data + nota na coluna Cliente, sem horas.

A IMPORTAÇÃO do histórico (fase 2, #125) vive aqui também: import_workbook()
lê as abas mensais da planilha real e converte em entries com origin='import'.
Autoridade de data: o NOME da aba manda no mês/ano (a planilha real tem abas
copiadas com ano errado nas células); da coluna A só sai o DIA.

Dependência: openpyxl (puro Python; a stdlib não escreve xlsx) — a única lib
nova do épico, usada para escrever (export) e ler (import).
"""

from __future__ import annotations

import logging
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import timesheet_db

log = logging.getLogger("scriba.timesheet_xlsx")

# -- layout (posições/fórmulas num lugar só; ajustar aqui se a planilha mudar) --

_HEADERS = ("Data", "Início", "Fim", "Total", "Cliente", "Projeto", "Descrição",
            "Responsavel", "Local", None, "Total dia", "Ñ Apontado ", "Apontados")
(_C_DATA, _C_INICIO, _C_FIM, _C_TOTAL, _C_CLIENTE, _C_PROJETO, _C_DESCRICAO,
 _C_RESP, _C_LOCAL, _C_APONTADO, _C_TOTAL_DIA, _C_NAO_APONT, _C_APONTADOS) = range(1, 14)

_OVERTIME_MARK = "hora extra"          # texto dominante da coluna H na planilha real
_FMT_DATA = "d-mmm"
_FMT_HORA = "h:mm"
_FMT_DUR = "h:mm"                      # D e K, como no original (blocos/dias < 24 h)
_FMT_DUR_MES = "[h]:mm"                # L/M: totais de mês passam de 24 h
_DV_LOCAL = '"Base,In Loco,Remoto"'
_DV_SIM_NAO = '"SIM,NÃO"'
_WIDTHS = {_C_DATA: 8, _C_INICIO: 7, _C_FIM: 7, _C_TOTAL: 7, _C_CLIENTE: 16,
           _C_PROJETO: 14, _C_DESCRICAO: 55, _C_RESP: 12, _C_LOCAL: 9,
           _C_APONTADO: 6, _C_TOTAL_DIA: 9, _C_NAO_APONT: 11, _C_APONTADOS: 11}


def _default_export_dir() -> Path:
    from . import config  # lazy: config.load() cria APP_DIR/config.toml

    ts = config.load().timesheet
    if ts.export_dir:
        return Path(ts.export_dir).expanduser()
    return Path.home() / "Documents" / "ScribaDev" / "Apontamentos"


def _unique_path(path: Path) -> Path:
    """Nunca sobrescreve: Apontamento_X.xlsx -> _2, _3... se já existir."""
    if not path.exists():
        return path
    for n in range(2, 1000):
        cand = path.with_stem(f"{path.stem}_{n}")
        if not cand.exists():
            return cand
    raise FileExistsError(f"sem nome livre para {path}")


def _hhmm_to_time(hhmm: str):
    return datetime.strptime(hhmm, "%H:%M").time()


def export_month(month: str, dest: Path | None = None) -> Path:
    """Exporta os apontamentos confirmados de `month` ('AAAA-MM') para um xlsx novo.

    `dest` é a PASTA de destino (None = [timesheet].export_dir do config, vazio =
    Documentos/ScribaDev/Apontamentos). Devolve o caminho gerado. ValueError se o
    mês é inválido ou não tem nada confirmado (nem dia especial) para exportar.
    """
    try:
        datetime.strptime(month, "%Y-%m")
    except ValueError:
        raise ValueError(f"mês inválido (use AAAA-MM): {month!r}") from None
    entries = timesheet_db.list_entries(month=month, status="confirmed")
    notes = timesheet_db.day_notes_month(month)
    if not entries and not notes:
        raise ValueError(f"nenhum apontamento confirmado em {month}")

    dest_dir = Path(dest) if dest is not None else _default_export_dir()
    dest_dir.mkdir(parents=True, exist_ok=True)
    out = _unique_path(dest_dir / f"Apontamento_{month}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = month
    bold = Font(bold=True)
    for col, header in enumerate(_HEADERS, start=1):
        if header:
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold
        ws.column_dimensions[get_column_letter(col)].width = _WIDTHS[col]
    ws.freeze_panes = "A2"

    # agrupa por dia (entries já vêm ordenadas por dia/hora); dias especiais juntos
    by_day: dict[str, list[dict]] = {}
    for e in entries:
        by_day.setdefault(e["work_date"], []).append(e)

    row = 2
    for day in sorted(set(by_day) | set(notes)):
        day_val = date.fromisoformat(day)
        first = row
        for e in by_day.get(day, ()):
            ws.cell(row=row, column=_C_DATA, value=day_val).number_format = _FMT_DATA
            ws.cell(row=row, column=_C_INICIO,
                    value=_hhmm_to_time(e["start_time"])).number_format = _FMT_HORA
            ws.cell(row=row, column=_C_FIM,
                    value=_hhmm_to_time(e["end_time"])).number_format = _FMT_HORA
            ws.cell(row=row, column=_C_TOTAL,
                    value=f"=C{row}-B{row}").number_format = _FMT_DUR
            ws.cell(row=row, column=_C_CLIENTE,
                    value=e["client_name"] or e["client_text"] or None)
            ws.cell(row=row, column=_C_PROJETO,
                    value=e["project_code"] or e["project_text"] or None)
            ws.cell(row=row, column=_C_DESCRICAO, value=e["description"] or None)
            if e["overtime"]:
                ws.cell(row=row, column=_C_RESP, value=_OVERTIME_MARK)
            if e["location"]:
                ws.cell(row=row, column=_C_LOCAL, value=e["location"])
            ws.cell(row=row, column=_C_APONTADO,
                    value="SIM" if e["posted"] else "NÃO")
            row += 1
        if day in by_day:  # Total dia na última linha do dia (dias de 1, 2 ou N blocos)
            ws.cell(row=row - 1, column=_C_TOTAL_DIA,
                    value=f"=SUM(D{first}:D{row - 1})").number_format = _FMT_DUR
        if day in notes:   # feriado/férias: linha só com a data e a nota
            ws.cell(row=row, column=_C_DATA, value=day_val).number_format = _FMT_DATA
            ws.cell(row=row, column=_C_CLIENTE, value=notes[day])
            row += 1

    last = row - 1
    ws.cell(row=2, column=_C_NAO_APONT,
            value='=SUMIF(J:J,"NÃO",D:D)').number_format = _FMT_DUR_MES
    ws.cell(row=2, column=_C_APONTADOS,
            value='=SUMIF(J:J,"SIM",D:D)').number_format = _FMT_DUR_MES

    dv_local = DataValidation(type="list", formula1=_DV_LOCAL, allow_blank=True)
    dv_apontado = DataValidation(type="list", formula1=_DV_SIM_NAO, allow_blank=True)
    ws.add_data_validation(dv_local)
    ws.add_data_validation(dv_apontado)
    col_i, col_j = get_column_letter(_C_LOCAL), get_column_letter(_C_APONTADO)
    dv_local.add(f"{col_i}2:{col_i}{last}")
    dv_apontado.add(f"{col_j}2:{col_j}{last}")

    wb.save(out)
    log.info("timesheet: mês %s exportado para %s", month, out)
    return out


# -- importação do histórico (fase 2, #125) ------------------------------------

_MONTH_NAMES = {"janeiro": 1, "fevereiro": 2, "marco": 3, "abril": 4, "maio": 5,
                "junho": 6, "julho": 7, "agosto": 8, "setembro": 9,
                "outubro": 10, "novembro": 11, "dezembro": 12}


@dataclass
class ImportReport:
    """Relatório do import - é ele que o --dry-run imprime e que alimenta o
    cadastro de aliases (grafias cruas de cliente com contagem)."""
    source: str
    dry_run: bool
    months: list[tuple[str, str, int]] = field(default_factory=list)  # (aba, mês, importadas)
    imported: int = 0
    duplicates: int = 0        # chave natural já existia (reimport = no-op)
    day_notes: int = 0
    fixed_dates: int = 0       # células com mês/ano divergente da aba (aba manda)
    ignored: list[tuple[str, int, str, str]] = field(default_factory=list)
    unresolved: Counter = field(default_factory=Counter)   # grafia crua -> contagem
    h_reaproveitada: Counter = field(default_factory=Counter)  # col H que não é hora extra
    skipped_sheets: list[str] = field(default_factory=list)

    def summary(self) -> str:
        head = "DRY-RUN (nada gravado)" if self.dry_run else "importado"
        out = [f"{self.source} - {head}:"]
        for sheet, month, n in self.months:
            out.append(f"  {month}  {n:>4} apontamento(s)  (aba {sheet!r})")
        out.append(f"total: {self.imported} importado(s), {self.duplicates} já "
                   f"existia(m), {self.day_notes} dia(s) especial(is), "
                   f"{len(self.ignored)} linha(s) ignorada(s)")
        if self.fixed_dates:
            out.append(f"datas com mês/ano divergente da aba (a aba manda): "
                       f"{self.fixed_dates}")
        if self.unresolved:
            out.append("clientes não resolvidos (cadastre/aliase e o histórico "
                       "unifica):")
            for raw, n in self.unresolved.most_common():
                out.append(f"  {n:>4}x  {raw}")
        if self.h_reaproveitada:
            out.append("coluna H ignorada (não é marca de hora extra): " + ", ".join(
                f"{v!r} ({n}x)" for v, n in self.h_reaproveitada.most_common()))
        if self.ignored:
            out.append("linhas ignoradas:")
            for sheet, line, reason, preview in self.ignored:
                out.append(f"  {sheet!r} linha {line}: {reason}  [{preview}]")
        if self.skipped_sheets:
            out.append("abas puladas (não são mês): " +
                       ", ".join(repr(s) for s in self.skipped_sheets))
        return "\n".join(out)


def _clean(value) -> str:
    """Célula -> texto normalizado (apara e colapsa espaços, inclusive NBSP -
    a planilha real tem '401442\\xa0' e '\\xa0GAP\\xa04.1.03-G0')."""
    if value is None:
        return ""
    return " ".join(str(value).split())


def _cell_time(value) -> str | None:
    """Célula de horário -> 'HH:MM' (time, datetime ou texto); None se não é hora."""
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}"
    if isinstance(value, datetime):
        return f"{value.hour:02d}:{value.minute:02d}"
    text = _clean(value)
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            t = datetime.strptime(text, fmt)
            return f"{t.hour:02d}:{t.minute:02d}"
        except ValueError:
            pass
    return None


def _sheet_month(name: str, year_votes: list[int]) -> str | None:
    """Nome da aba -> 'AAAA-MM'; None se a aba não é um mês (ex.: 'Modelo (3)').

    O mês vem do NOME (acento e caixa-insensível: 'março(25)', 'Julho ( 26)');
    o ano, do sufixo numérico (25 -> 2025). Aba sem sufixo (os meses de 2024)
    usa a MAIORIA dos anos das datas da própria aba - não confiar em célula
    individual: abas copiadas carregam datas com ano errado.
    """
    folded = timesheet_db._fold(name)
    month = next((n for token, n in _MONTH_NAMES.items() if token in folded), None)
    if month is None:
        return None
    m = re.search(r"\d{4}|\d{2}", folded)
    if m:
        year = int(m.group())
        year = year if year >= 2000 else 2000 + year
    elif year_votes:
        year = Counter(year_votes).most_common(1)[0][0]
    else:
        return None
    return f"{year:04d}-{month:02d}"


def import_workbook(path: Path | str, dry_run: bool = False) -> ImportReport:
    """Importa o histórico da planilha de apontamento (todas as abas mensais).

    Cada linha com horário vira um entry origin='import' status='confirmed'
    (histórico é fato consumado), posted pela coluna J (SIM/NÃO); linha de dia
    especial (data + texto no Cliente, sem horas: 'feriado'...) vira day_note.
    Idempotente pela chave natural (dia, início, fim, cliente _fold): reimportar
    - ou importar por cima do que o app já registrou - não duplica.
    Com dry_run, NADA é gravado: só o relatório.
    """
    path = Path(path)
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        raise ValueError(f"não consegui abrir {path.name!r} — feche o arquivo no "
                         "Excel e tente de novo") from None
    report = ImportReport(source=str(path), dry_run=dry_run)

    # chaves naturais do que JÁ existe (qualquer origem; descartada não conta)
    existing = {
        (e["work_date"], e["start_time"], e["end_time"],
         timesheet_db._fold(e["client_name"] or e["client_text"] or ""))
        for e in timesheet_db.list_entries() if e["status"] != "discarded"
    }
    resolved: dict[str, tuple[int | None, str]] = {}   # cache: _fold -> (id, canônico)

    try:
        for sheet in wb.sheetnames:
            rows = [(r + (None,) * 10)[:10] for r in
                    wb[sheet].iter_rows(min_row=1, max_col=10, values_only=True)]
            votes = [a.year for a, *_ in rows if isinstance(a, (datetime, date))]
            month = _sheet_month(sheet, votes)
            if month is None:
                report.skipped_sheets.append(sheet)
                continue
            n0 = report.imported
            _import_sheet(sheet, month, rows, report, existing, resolved, dry_run)
            report.months.append((sheet, month, report.imported - n0))
    finally:
        wb.close()
    log.info("timesheet: import de %s -> %d novos, %d duplicados, %d ignorados%s",
             path, report.imported, report.duplicates, len(report.ignored),
             " (dry-run)" if dry_run else "")
    return report


def _import_sheet(sheet: str, month: str, rows: list, report: ImportReport,
                  existing: set, resolved: dict, dry_run: bool) -> None:
    year, month_num = int(month[:4]), int(month[5:7])
    current_day: str | None = None

    for line, row in enumerate(rows, start=1):
        a, b, c, _total, e, f, g, h, loc, j = row
        cliente, projeto, desc = _clean(e), _clean(f), _clean(g)
        start, end = _cell_time(b), _cell_time(c)
        if cliente == "Cliente" or _clean(a) == "Data":
            continue                       # cabeçalho
        preview = " | ".join(x for x in (cliente, projeto, desc) if x)[:60]

        if a is not None and _clean(a):    # célula de data abre um dia novo
            if isinstance(a, (datetime, date)):
                if (a.month, a.year) != (month_num, year):
                    report.fixed_dates += 1      # aba manda; da célula sai só o DIA
                day = a.day
            else:
                m = re.fullmatch(r"(\d{1,2})([/-]\d{1,2}([/-]\d{2,4})?)?",
                                 _clean(a))
                day = int(m.group(1)) if m else None
            try:
                current_day = date(year, month_num, day).isoformat() if day else None
            except (ValueError, TypeError):
                current_day = None
            if current_day is None:
                report.ignored.append((sheet, line, f"data ilegível: {a!r}", preview))
                continue

        if not (cliente or desc or start or end):
            continue                       # sobra do template (só D=0:00 e J)

        if start is None and end is None:
            if current_day and cliente and a is not None:
                # dia especial na própria linha da data: feriado, férias...
                note = cliente + (f" - {desc}" if desc else "")
                if not dry_run:
                    timesheet_db.set_day_note(current_day, note)
                report.day_notes += 1
            else:
                report.ignored.append((sheet, line, "sem horas", preview))
            continue
        if start is None or end is None or current_day is None:
            reason = "sem data válida" if current_day is None else "horário incompleto"
            report.ignored.append((sheet, line, reason, preview))
            continue
        if end <= start:
            report.ignored.append(
                (sheet, line, f"fim ({end}) não é depois do início ({start})", preview))
            continue

        overtime = "extra" in timesheet_db._fold(h or "")
        h_txt = _clean(h)
        if h_txt and not overtime and h_txt != "Responsavel":
            report.h_reaproveitada[h_txt] += 1

        ckey = timesheet_db._fold(cliente)
        if ckey not in resolved:
            resolved[ckey] = timesheet_db.resolve_client(cliente)
        cid, cname = resolved[ckey]     # canônico quando resolve; senão o cru
        if cid is None and cliente:     # grafias fold-iguais contam juntas
            report.unresolved[cname] += 1

        # chave natural pelo nome que fica no banco (canônico p/ resolvido):
        # 'Usina Vetra' via alias deduplica contra 'Vetra' já gravado
        key = (current_day, start, end, timesheet_db._fold(cname))
        if key in existing:
            report.duplicates += 1
            continue
        existing.add(key)                  # dedupe também DENTRO da planilha

        pid, ptext = None, projeto
        if cid is not None and projeto:
            match = [p for p in timesheet_db.list_projects(cid)
                     if p["code"].casefold() == projeto.casefold()]
            if match:
                pid, ptext = match[0]["id"], ""
        if not dry_run:
            timesheet_db.add_entry(
                work_date=current_day, start_time=start, end_time=end,
                client_id=cid, client_text="" if cid is not None else cname,
                project_id=pid, project_text=ptext, description=desc,
                overtime=overtime, location=_clean(loc),
                posted=timesheet_db._fold(j or "") == "sim", origin="import")
        report.imported += 1
