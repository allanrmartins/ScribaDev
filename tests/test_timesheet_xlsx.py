"""Testes da exportação e importação Excel do timesheet (épico #118, #122/#125).

Exporta um mês sintético (feriado + dias com 1, 2 e 3 blocos) e RELÊ com
openpyxl validando: layout A-M (cabeçalhos exatos da planilha real, incluindo
'Ñ Apontado ' com espaço), tipos reais (date/time) e formatos, fórmulas D/K
(Total dia generalizado para N blocos), L/M de resumo, marca 'hora extra' na H,
SIM/NÃO refletindo `posted`, dropdowns (data validation), exclusão de
suggested/discarded e não-sobrescrita (sufixo _2).

A importação (#125) é testada com uma planilha sintética que REPRODUZ as
pegadinhas da real: aba com ano só no nome, células de data com ano errado
(aba copiada), NBSP em projeto, coluna H reaproveitada, linhas de anotação sem
horas, sobras de template (D=0:00 + J) e aba 'Modelo'.

Roda sem dependências externas além do openpyxl (dependência do projeto):
python -m unittest discover -s tests
"""

import shutil
import sys
import tempfile
import unittest
from datetime import datetime, time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from scriba import timesheet_db as tsdb  # noqa: E402
from scriba import timesheet_xlsx as tsx  # noqa: E402
from scriba import util  # noqa: E402


class TimesheetXlsxTests(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="scriba_tsx_"))
        self._app0, self._logs0, self._db0 = util.APP_DIR, util.LOGS_DIR, tsdb.DB_PATH
        util.APP_DIR = self.tmp / "app"
        util.LOGS_DIR = util.APP_DIR / "logs"
        tsdb.DB_PATH = self.tmp / "timesheet.db"
        self.out_dir = self.tmp / "out"
        self._seed()

    def tearDown(self):
        util.APP_DIR, util.LOGS_DIR, tsdb.DB_PATH = self._app0, self._logs0, self._db0
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _seed(self):
        cid = tsdb.add_client("Coruripe")
        pid = tsdb.add_project(cid, "403240", "Reforma tributária")
        tsdb.set_day_note("2026-07-09", "feriado")
        # dia 13: 2 blocos (manhã apontada; tarde extra, não apontada, cliente cru)
        e1 = tsdb.add_entry(work_date="2026-07-13", start_time="08:00", end_time="13:00",
                            client_id=cid, project_id=pid,
                            description="aplicacao de notas outbound")
        tsdb.set_posted([e1], True)
        tsdb.add_entry(work_date="2026-07-13", start_time="14:00", end_time="17:00",
                       client_text="Delta", description="ajuste MM03", overtime=True)
        # dia 14: 1 bloco
        tsdb.add_entry(work_date="2026-07-14", start_time="08:00", end_time="12:00",
                       client_id=cid, description="testes unitarios")
        # dia 15: 3 blocos (generalização do Total dia)
        for start, end in (("08:00", "10:00"), ("10:00", "12:00"), ("14:00", "16:00")):
            tsdb.add_entry(work_date="2026-07-15", start_time=start, end_time=end,
                           client_id=cid, description="call checkpoint")
        # não exportáveis: sugestão pendente e descartada
        tsdb.upsert_suggestion({"work_date": "2026-07-13", "start_time": "18:00",
                                "end_time": "19:00", "description": "sugerida",
                                "meeting_started_at": "2026-07-13T18:00:00"})
        tsdb.add_entry(work_date="2026-07-14", start_time="18:00", end_time="19:00",
                       status="discarded", description="descartada")

    def _export(self):
        return tsx.export_month("2026-07", self.out_dir)

    # -- layout e conteúdo ------------------------------------------------------
    def test_layout_completo(self):
        out = self._export()
        self.assertEqual(out.name, "Apontamento_2026-07.xlsx")
        ws = load_workbook(out)["2026-07"]

        # cabeçalhos exatos da planilha real (J sem cabeçalho; L com espaço final)
        headers = [ws.cell(row=1, column=c).value for c in range(1, 14)]
        self.assertEqual(headers, list(tsx._HEADERS))

        # linhas: 2=feriado(09), 3-4=dia13, 5=dia14, 6-8=dia15; nada além disso
        self.assertEqual(ws.max_row, 8)

        # feriado: só data + nota na coluna Cliente, sem horas
        self.assertEqual(str(ws["A2"].value)[:10], "2026-07-09")
        self.assertEqual(ws["E2"].value, "feriado")
        self.assertIsNone(ws["B2"].value)
        self.assertIsNone(ws["K2"].value)

        # tipos reais + formatos (como na planilha original)
        self.assertEqual(str(ws["A3"].value)[:10], "2026-07-13")
        self.assertEqual(ws["A3"].number_format, "d-mmm")
        self.assertEqual(ws["B3"].value, time(8, 0))
        self.assertEqual(ws["B3"].number_format, "h:mm")
        self.assertEqual(ws["D3"].value, "=C3-B3")
        self.assertEqual(ws["D3"].number_format, "h:mm")

        # cliente canônico + projeto cadastrado; cliente cru preservado
        self.assertEqual(ws["E3"].value, "Coruripe")
        self.assertEqual(ws["F3"].value, "403240")
        self.assertEqual(ws["E4"].value, "Delta")

        # hora extra na H (texto real da planilha); SIM/NÃO refletindo posted
        self.assertIsNone(ws["H3"].value)
        self.assertEqual(ws["H4"].value, "hora extra")
        self.assertEqual(ws["J3"].value, "SIM")
        self.assertEqual(ws["J4"].value, "NÃO")

        # Total dia na ÚLTIMA linha de cada dia: 2, 1 e 3 blocos
        self.assertIsNone(ws["K3"].value)
        self.assertEqual(ws["K4"].value, "=SUM(D3:D4)")
        self.assertEqual(ws["K5"].value, "=SUM(D5:D5)")
        self.assertEqual(ws["K8"].value, "=SUM(D6:D8)")

        # resumo do mês (L/M) com formato de duração longa
        self.assertEqual(ws["L2"].value, '=SUMIF(J:J,"NÃO",D:D)')
        self.assertEqual(ws["M2"].value, '=SUMIF(J:J,"SIM",D:D)')
        self.assertEqual(ws["L2"].number_format, "[h]:mm")

        # a sugestão (18:00) e a descartada não entram
        contents = [ws.cell(row=r, column=7).value for r in range(2, 9)]
        self.assertNotIn("sugerida", contents)
        self.assertNotIn("descartada", contents)

    def test_dropdowns(self):
        ws = load_workbook(self._export())["2026-07"]
        dvs = ws.data_validations.dataValidation
        self.assertEqual(len(dvs), 2)
        formulas = {dv.formula1 for dv in dvs}
        self.assertIn('"SIM,NÃO"', formulas)
        self.assertIn('"Base,In Loco,Remoto"', formulas)
        ranges = {str(dv.sqref) for dv in dvs}
        self.assertIn("J2:J8", ranges)
        self.assertIn("I2:I8", ranges)

    # -- não-sobrescrita e erros --------------------------------------------------
    def test_nunca_sobrescreve(self):
        first = self._export()
        second = self._export()
        self.assertEqual(second.name, "Apontamento_2026-07_2.xlsx")
        self.assertTrue(first.exists())
        self.assertTrue(second.exists())

    def test_mes_vazio_e_invalido(self):
        with self.assertRaises(ValueError):
            tsx.export_month("2031-01", self.out_dir)
        with self.assertRaises(ValueError):
            tsx.export_month("07/2026", self.out_dir)

    def test_mes_so_com_feriado_exporta(self):
        tsdb.set_day_note("2026-08-07", "férias")
        ws = load_workbook(tsx.export_month("2026-08", self.out_dir))["2026-08"]
        self.assertEqual(ws.max_row, 2)
        self.assertEqual(ws["E2"].value, "férias")


class TimesheetImportTests(unittest.TestCase):
    """Importação do histórico (#125) sobre planilha com as pegadinhas reais."""

    _HEAD = ["Data", "Início", "Fim", "Total", "Cliente", "Projeto",
             "Descrição", "Responsavel", "Local", None]

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="scriba_tsi_"))
        self._app0, self._logs0, self._db0 = util.APP_DIR, util.LOGS_DIR, tsdb.DB_PATH
        util.APP_DIR = self.tmp / "app"
        util.LOGS_DIR = util.APP_DIR / "logs"
        tsdb.DB_PATH = self.tmp / "timesheet.db"

    def tearDown(self):
        util.APP_DIR, util.LOGS_DIR, tsdb.DB_PATH = self._app0, self._logs0, self._db0
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _planilha(self) -> Path:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Julho ( 26)"           # nome de aba irregular da planilha real
        ws.append(self._HEAD)
        # célula com ANO ERRADO (aba copiada): o nome da aba manda -> 2026-07-01
        ws.append([datetime(2025, 7, 1), time(8, 0), time(13, 0), time(5, 0),
                   "Coruripe", "403240\xa0", "aplicacao de notas", None, None, "SIM"])
        ws.append([None, time(14, 0), time(17, 0), time(3, 0), "coruripe",
                   "403240", "aplicacao de notas", "hora extra", None, "NÃO"])
        ws.append([None, None, None, time(0, 0), None, None, None, None, None,
                   "NÃO"])                 # sobra do template: some em silêncio
        ws.append([datetime(2026, 7, 2), None, None, time(0, 0), "feriado",
                   None, None, None, None, "NÃO"])   # dia especial -> day_note
        ws.append([datetime(2026, 7, 3), time(8, 0), time(12, 0), time(4, 0),
                   "Usina Coruripe", None, "call kickoff",
                   "solicitado monique", None, None])  # H reaproveitada, J vazio
        ws.append([None, None, None, time(0, 0), "Simpar", "FM / WF",
                   "anotacao sem horas", None, None, "NÃO"])  # ignorada c/ motivo
        ws.append([None, time(15, 0), time(14, 0), None, "Delta", None,
                   "fim antes do inicio", None, None, "NÃO"])  # ignorada c/ motivo
        ws2 = wb.create_sheet("agosto")    # SEM ano no nome: maioria das datas
        ws2.append(self._HEAD)
        ws2.append([datetime(2026, 8, 3), time(9, 0), time(11, 0), None,
                    "Célera", "\xa0GAP\xa04.1.03-G01", "dev relatorio", None,
                    "Remoto", "SIM"])
        wb.create_sheet("Modelo (3)").append(self._HEAD)   # não é mês: pulada
        path = self.tmp / "Apontamento.xlsx"
        wb.save(path)
        return path

    def test_import_completo(self):
        report = tsx.import_workbook(self._planilha())
        self.assertEqual(report.imported, 4)
        self.assertEqual(report.duplicates, 0)
        self.assertEqual(report.day_notes, 1)
        self.assertEqual(report.fixed_dates, 1)       # 2025-07-01 na aba de 2026
        self.assertEqual(report.skipped_sheets, ["Modelo (3)"])
        self.assertEqual([m[1] for m in report.months], ["2026-07", "2026-08"])
        self.assertEqual(sorted(r[2] for r in report.ignored),
                         ["fim (14:00) não é depois do início (15:00)", "sem horas"])
        # grafias fold-iguais contam juntas; Célera com acento preservado
        self.assertEqual(report.unresolved,
                         {"Coruripe": 2, "Usina Coruripe": 1, "Célera": 1})
        self.assertIn("solicitado monique", report.h_reaproveitada)

        # o que caiu no banco: dia com ano corrigido, NBSP normalizado, posted
        # da coluna J, hora extra da H, origin='import'
        dia1 = tsdb.list_entries(day="2026-07-01")
        self.assertEqual(len(dia1), 2)
        self.assertEqual(dia1[0]["project_text"], "403240")
        self.assertEqual((dia1[0]["posted"], dia1[1]["posted"]), (1, 0))
        self.assertIsNotNone(dia1[0]["posted_at"])
        self.assertEqual(dia1[1]["overtime"], 1)
        self.assertTrue(all(e["origin"] == "import" and e["status"] == "confirmed"
                            for e in dia1))
        self.assertEqual(tsdb.day_notes_month("2026-07"), {"2026-07-02": "feriado"})
        ago = tsdb.list_entries(day="2026-08-03")[0]
        self.assertEqual(ago["project_text"], "GAP 4.1.03-G01")
        self.assertEqual(ago["location"], "Remoto")

    def test_dry_run_nao_grava_e_reimport_nao_duplica(self):
        path = self._planilha()
        dry = tsx.import_workbook(path, dry_run=True)
        self.assertTrue(dry.dry_run)
        self.assertEqual(dry.imported, 4)
        self.assertEqual(tsdb.list_entries(), [])                 # nada gravado
        self.assertEqual(tsdb.day_notes_month("2026-07"), {})

        tsx.import_workbook(path)
        again = tsx.import_workbook(path)                         # reimport: no-op
        self.assertEqual(again.imported, 0)
        self.assertEqual(again.duplicates, 4)
        self.assertEqual(len(tsdb.list_entries()), 4)

    def test_resolucao_e_dedupe_contra_o_banco(self):
        # cliente cadastrado + alias: import resolve (acento-insensível) e a
        # chave natural deduplica contra o que o APP já registrou no mesmo slot
        cid = tsdb.add_client("Coruripe")
        tsdb.add_alias(cid, "usina coruripe")
        tsdb.add_client("Celera")           # sem acento: casa "Célera" via _fold
        tsdb.add_entry(work_date="2026-07-01", start_time="08:00",
                       end_time="13:00", client_id=cid,
                       description="registrado pelo app")
        report = tsx.import_workbook(self._planilha())
        self.assertEqual(report.unresolved, {})                   # tudo resolvido
        self.assertEqual(report.duplicates, 1)                    # slot já existia
        self.assertEqual(report.imported, 3)
        kick = tsdb.list_entries(day="2026-07-03")[0]
        self.assertEqual(kick["client_id"], cid)                  # via alias
        self.assertEqual(kick["client_name"], "Coruripe")

    def test_planilha_inexistente_ou_travada(self):
        with self.assertRaises(Exception):
            tsx.import_workbook(self.tmp / "nao-existe.xlsx")


if __name__ == "__main__":
    unittest.main()
